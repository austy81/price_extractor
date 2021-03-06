from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, colors
import time
import logging


class excel:

    def __init__(self, input_xslx_file, output_xslx_file, insert_sheet_name):
        self.in_wb = load_workbook(input_xslx_file)
        self.output_xslx_file = output_xslx_file
        self.insert_sheet_name = "{}-{}".format(
            insert_sheet_name, time.strftime("%Y %d. %m. %H.%M.%S"))

    def get_column_values(self, sheet_name, url_column, skip_first_line=True):
        sheet = self.in_wb[sheet_name]
        urls = []
        for row in sheet:
            if row[0].row == 1 and skip_first_line:
                continue
            if row[url_column].value is None:
                break
            urls.append(
                {"row_number": row[url_column].row, "url": row[url_column].value})
        return urls

    def get_parsers(self, sheet_name, skip_first_line=True):
        ws = self.in_wb[sheet_name]
        parsers = []
        for row in ws:
            if row[0].row == 1 and skip_first_line:
                continue
            if row[0].value is None or row[0].value == "":
                break
            row_no = str(row[0].row)
            url_cell = ws['A' + row_no]

            # this is fuck up in openpyxl
            bg = url_cell.fill.fgColor.rgb if url_cell.fill.bgColor.rgb != colors.BLACK else colors.WHITE

            parsers.append(
                {'url_regex': url_cell.value,
                 'price_element': ws['B' + row_no].value,
                 'verify_exists': ws['C' + row_no].value,
                 'verify_not_exists': ws['D' + row_no].value,
                 'parser_cell_fill_bg_color': bg,
                 'parser_name': ws['E' + row_no].value})
        return parsers

    def save_in_excel(self, results):
        if len(results) == 0:
            logging.warning("Nothing to save.")
            return
        ws = self._load_out_worksheet()
        ws['A1'].value = "URL"
        ws['B1'].value = "Price"
        ws['C1'].value = "Error"
        for parser_urls in results:
            parser = parser_urls['parser']
            for row in parser_urls["urls"]:
                error_text, result_color = row['note'] if row['note'] else (None, None)
                priece_cell = ws['B'+str(row["row_number"])]

                ws.cell(row=row["row_number"], column=1, value=row["url"])
                try:
                    priece_cell.fill = PatternFill("solid", result_color) if result_color else PatternFill(
                        "solid", fgColor=parser['parser_cell_fill_bg_color'])
                    priece_cell.value = row['price'] if row['price'] else 'n/a'
                except Exception as e:
                    ws.cell(row=row["row_number"], column=2, value=repr(e))
                ws.cell(row=row["row_number"], column=3, value=error_text)

        self._save_workbook()

    def _get_fill_color(self, parser_color, result_color):
        return PatternFill("solid", result_color) if result_color else PatternFill("solid", fgColor=parser_color)

    def _save_workbook(self):
        while True:
            try:
                self.out_wb.save(self.output_xslx_file)
                break
            except Exception:
                logging.info('Save to file {} failed. Retrying...'.format(
                            self.output_xslx_file))
                time.sleep(5)
        logging.info("Output excel file {} was saved.".format(
            self.output_xslx_file))

    def _load_out_worksheet(self):
        try:
            self.out_wb = load_workbook(self.output_xslx_file)
            logging.info("Output excel loaded.")
        except Exception:
            logging.warn("Excel workbook {} was not found.".format(
                self.output_xslx_file))
            logging.info("Creating new workbook.")
            self.out_wb = Workbook()
        self.out_wb.create_sheet(title=self.insert_sheet_name)
        return self.out_wb[self.insert_sheet_name]
